from __future__ import annotations

import csv
import re
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "calendar_chm_2026_msk.xlsx"
MSK = ZoneInfo("Europe/Moscow")


RU_TEAMS = {
    "Mexico": "Мексика",
    "South Africa": "ЮАР",
    "South Korea": "Южная Корея",
    "Czechia": "Чехия",
    "Canada": "Канада",
    "Bosnia & Herzegovina": "Босния и Герцеговина",
    "USA": "США",
    "Paraguay": "Парагвай",
    "Qatar": "Катар",
    "Switzerland": "Швейцария",
    "Brazil": "Бразилия",
    "Morocco": "Марокко",
    "Haiti": "Гаити",
    "Scotland": "Шотландия",
    "Australia": "Австралия",
    "Türkiye": "Турция",
    "Germany": "Германия",
    "Curaçao": "Кюрасао",
    "Netherlands": "Нидерланды",
    "Japan": "Япония",
    "Ivory Coast": "Кот-д’Ивуар",
    "Ecuador": "Эквадор",
    "Sweden": "Швеция",
    "Tunisia": "Тунис",
    "Spain": "Испания",
    "Cape Verde": "Кабо-Верде",
    "Belgium": "Бельгия",
    "Egypt": "Египет",
    "Saudi Arabia": "Саудовская Аравия",
    "Uruguay": "Уругвай",
    "Iran": "Иран",
    "New Zealand": "Новая Зеландия",
    "France": "Франция",
    "Senegal": "Сенегал",
    "Iraq": "Ирак",
    "Norway": "Норвегия",
    "Argentina": "Аргентина",
    "Algeria": "Алжир",
    "Austria": "Австрия",
    "Jordan": "Иордания",
    "Portugal": "Португалия",
    "DR Congo": "ДР Конго",
    "England": "Англия",
    "Croatia": "Хорватия",
    "Ghana": "Гана",
    "Panama": "Панама",
    "Uzbekistan": "Узбекистан",
    "Colombia": "Колумбия",
}

STAGE_LABELS = {
    "RO32": "1/16 финала",
    "RO16": "1/8 финала",
    "QF": "Четвертьфинал",
    "SF": "Полуфинал",
    "3RD": "Матч за 3 место",
    "FIN": "Финал",
}


def ru_side(name: str) -> str:
    if name in RU_TEAMS:
        return RU_TEAMS[name]
    name = re.sub(r"^Runner-up ([A-L])$", r"2-е место группы \1", name)
    name = re.sub(r"^Winner ([A-L])$", r"Победитель группы \1", name)
    name = re.sub(r"^Best 3rd \((.+)\)$", r"Лучшая 3-я команда (\1)", name)
    name = re.sub(r"^W(\d+)$", r"Победитель матча \1", name)
    name = re.sub(r"^L(\d+)$", r"Проигравший матча \1", name)
    return name


def ru_match(row: dict[str, str]) -> str:
    return f"{ru_side(row['team_a'])} - {ru_side(row['team_b'])}"


def stage_label(code: str) -> str:
    return STAGE_LABELS.get(code, f"Группа {code}")


def load_rows() -> list[dict[str, str]]:
    src = (ROOT / "build_worldcup_calendar.mjs").read_text(encoding="utf-8")
    match = re.search(r"const csv = `(.+?)`;", src, flags=re.S)
    if not match:
        raise RuntimeError("CSV schedule block not found")
    return list(csv.DictReader(match.group(1).splitlines()))


def enrich(rows: list[dict[str, str]]) -> list[dict[str, object]]:
    out = []
    for row in rows:
        local_date = datetime.strptime(row["date"], "%Y-%m-%d").date()
        hour, minute = map(int, row["kickoff_local"].split(":"))
        local_dt = datetime(local_date.year, local_date.month, local_date.day, hour, minute, tzinfo=ZoneInfo(row["stadium_tz"]))
        msk_dt = local_dt.astimezone(MSK)
        row = dict(row)
        row["match_no"] = int(row["match_no"])
        row["match_ru"] = ru_match(row)
        row["stage_ru"] = stage_label(row["group"])
        row["msk_dt"] = msk_dt
        row["msk_date"] = msk_dt.date()
        row["msk_time"] = msk_dt.strftime("%H:%M")
        row["local_text"] = f"{row['date']} {row['kickoff_local']}"
        out.append(row)
    return sorted(out, key=lambda item: (item["msk_dt"], item["match_no"]))


def style_ws(ws):
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color="1F2933")


def set_title(ws, cell_range: str, title: str, subtitle: str | None = None):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = title
    cell.fill = PatternFill("solid", fgColor="12355B")
    cell.font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[cell.row].height = 32
    if subtitle:
        ws.cell(cell.row + 1, 1).value = subtitle
        ws.cell(cell.row + 1, 1).font = Font(name="Aptos", italic=True, color="667085")


def apply_table_style(ws, min_row: int, max_row: int, min_col: int, max_col: int):
    thin = Side(style="thin", color="D9E2EC")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for cell in ws[min_row]:
        if min_col <= cell.column <= max_col:
            cell.fill = PatternFill("solid", fgColor="2F6F9F")
            cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def monday_of(day):
    return day - timedelta(days=day.weekday())


def build_calendar(ws, matches):
    set_title(ws, "A1:H1", "ЧМ-2026: календарь матчей по московскому времени", "Дни недели сверху, время МСК слева, матчи внутри ячеек.")
    ws["A3"] = "Как читать"
    ws["B3"] = "Найди дату сверху, время слева, матч в пересечении."
    ws.merge_cells("B3:H3")
    for cell in ws[3]:
        cell.fill = PatternFill("solid", fgColor="EAF3F8")
        cell.font = Font(name="Aptos", bold=(cell.column == 1), color="1F2933")

    weekdays = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    week_start = monday_of(matches[0]["msk_date"])
    max_date = matches[-1]["msk_date"]
    row = 5
    thin = Side(style="thin", color="D9E2EC")

    while week_start <= max_date:
        days = [week_start + timedelta(days=i) for i in range(7)]
        week_matches = [m for m in matches if m["msk_date"] in days]
        times = sorted({m["msk_time"] for m in week_matches}) or [""]
        body_start = row + 3
        body_end = body_start + len(times) - 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row, 1).value = f"Неделя {week_start:%d.%m.%Y} - {days[-1]:%d.%m.%Y}"
        ws.cell(row, 1).fill = PatternFill("solid", fgColor="12355B")
        ws.cell(row, 1).font = Font(name="Aptos", bold=True, color="FFFFFF")
        ws.cell(row, 1).alignment = Alignment(horizontal="center")

        headers = ["Время МСК", *weekdays]
        dates = ["", *[f"{d:%d.%m.%Y}" for d in days]]
        for col, value in enumerate(headers, 1):
            c = ws.cell(row + 1, col, value)
            c.fill = PatternFill("solid", fgColor="2F6F9F")
            c.font = Font(name="Aptos", bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
        for col, value in enumerate(dates, 1):
            c = ws.cell(row + 2, col, value)
            c.fill = PatternFill("solid", fgColor="FFF4D8")
            c.font = Font(name="Aptos", bold=True)
            c.alignment = Alignment(horizontal="center")

        slot_map = {}
        for m in week_matches:
            slot_map.setdefault((m["msk_date"], m["msk_time"]), []).append(m)

        for idx, time in enumerate(times):
            r = body_start + idx
            ws.cell(r, 1, time)
            ws.cell(r, 1).fill = PatternFill("solid", fgColor="EAF3F8")
            ws.cell(r, 1).font = Font(name="Aptos", bold=True)
            ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[r].height = 44
            for col, day in enumerate(days, 2):
                items = slot_map.get((day, time), [])
                c = ws.cell(r, col)
                if items:
                    c.value = "\n\n".join(m["match_ru"] for m in items)
                    c.fill = PatternFill("solid", fgColor="EAF7F3" if all(len(m["group"]) == 1 for m in items) else "FFF4D8")
                    c.font = Font(name="Aptos", size=10, bold=True, color="1F2933")
                    if len(items) > 1:
                        ws.row_dimensions[r].height = 70
                elif col >= 7:
                    c.fill = PatternFill("solid", fgColor="F1F5F9")
                c.alignment = Alignment(vertical="top", wrap_text=True)

        for r in range(row, body_end + 1):
            for col in range(1, 9):
                ws.cell(r, col).border = Border(left=thin, right=thin, top=thin, bottom=thin)

        row = body_end + 3
        week_start += timedelta(days=7)

    widths = [12, 28, 28, 28, 28, 28, 28, 28]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A4"


def build_list(ws, matches):
    set_title(ws, "A1:K1", "Список матчей ЧМ-2026: время по Москве", "Здесь можно отмечать матчи для просмотра.")
    headers = ["Хочу смотреть", "Приоритет", "Дата МСК", "Время МСК", "Матч", "Стадия", "#", "Стадион", "Город", "Локальное время", "Заметки"]
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append(headers)
    for m in matches:
        ws.append(["", "", m["msk_date"], m["msk_time"], m["match_ru"], m["stage_ru"], m["match_no"], m["stadium"], m["city"], m["local_text"], ""])
    apply_table_style(ws, 5, 5 + len(matches), 1, len(headers))
    for row in range(6, 6 + len(matches)):
        ws.cell(row, 3).number_format = "dd.mm.yyyy"
    dv_watch = DataValidation(type="list", formula1='"Да,Нет"', allow_blank=True)
    dv_priority = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    ws.add_data_validation(dv_watch)
    ws.add_data_validation(dv_priority)
    dv_watch.add(f"A6:A{5 + len(matches)}")
    dv_priority.add(f"B6:B{5 + len(matches)}")
    ws.conditional_formatting.add(f"A6:K{5 + len(matches)}", FormulaRule(formula=['$A6="Да"'], fill=PatternFill("solid", fgColor="EAF7F3")))
    widths = [16, 12, 12, 11, 34, 16, 7, 24, 16, 18, 24]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.auto_filter.ref = f"A5:K{5 + len(matches)}"
    ws.freeze_panes = "A6"


def build_summary(ws, title_text, headers, rows, widths):
    set_title(ws, f"A1:{get_column_letter(len(headers))}1", title_text)
    ws.append([])
    ws.append([])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    apply_table_style(ws, 4, 4 + len(rows), 1, len(headers))
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def main():
    matches = enrich(load_rows())
    wb = Workbook()
    wb.remove(wb.active)

    ws_calendar = wb.create_sheet("Календарь")
    ws_list = wb.create_sheet("Список МСК")
    ws_pick = wb.create_sheet("Мой выбор")
    ws_groups = wb.create_sheet("Группы")
    ws_venues = wb.create_sheet("Стадионы")
    ws_source = wb.create_sheet("Источник")

    build_calendar(ws_calendar, matches)
    build_list(ws_list, matches)

    set_title(ws_pick, "A1:F1", "Мой выбор", "Отмеченные матчи удобнее смотреть через фильтр на листе «Список МСК»: колонка «Хочу смотреть» = Да.")
    ws_pick["A4"] = "Зачем этот лист"
    ws_pick["B4"] = "Чтобы не усложнять файл хрупкими формулами, выбранные матчи фильтруются на листе «Список МСК»."
    ws_pick.column_dimensions["A"].width = 18
    ws_pick.column_dimensions["B"].width = 90

    group_counts = {}
    for m in matches:
        group_counts.setdefault(m["group"], [m["stage_ru"], 0])
        group_counts[m["group"]][1] += 1
    order = ["A","B","C","D","E","F","G","H","I","J","K","L","RO32","RO16","QF","SF","3RD","FIN"]
    group_rows = [[g, group_counts[g][0], group_counts[g][1]] for g in order if g in group_counts]
    build_summary(ws_groups, "Группы и стадии", ["Код", "Стадия", "Матчей"], group_rows, [10, 20, 12])

    venue_counts = {}
    for m in matches:
        key = (m["stadium"], m["city"])
        venue_counts[key] = venue_counts.get(key, 0) + 1
    venue_rows = [[stadium, city, count] for (stadium, city), count in sorted(venue_counts.items(), key=lambda x: (-x[1], x[0][1]))]
    build_summary(ws_venues, "Стадионы", ["Стадион", "Город", "Матчей"], venue_rows, [28, 18, 12])

    set_title(ws_source, "A1:B1", "Источник и примечания")
    source_rows = [
        ["Основной источник", "https://kickofftimes.tv/"],
        ["Официальная сверка", "https://www.fifa.com/en/tournaments/mens/worldcup/canadamexicousa2026/articles/match-schedule-fixtures-results-teams-stadiums"],
        ["Важно", "Kickoff Times указывает, что это независимое фанатское расписание, не связанное с FIFA."],
        ["Часовой пояс файла", "Europe/Moscow / МСК UTC+3"],
        ["Дата сборки", datetime.now().strftime("%Y-%m-%d")],
    ]
    ws_source.append([])
    ws_source.append([])
    ws_source.append(["Что", "Значение"])
    for row in source_rows:
        ws_source.append(row)
    apply_table_style(ws_source, 4, 4 + len(source_rows), 1, 2)
    ws_source.column_dimensions["A"].width = 24
    ws_source.column_dimensions["B"].width = 110

    for ws in wb.worksheets:
        style_ws(ws)

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
